import os
from openpyxl import load_workbook
from scripts.handle_path import DATA_PATH


class Testcase:
    pass


class HandleExcel:

    def __init__(self, filename, sheetname=None):
        self.filename = os.path.join(DATA_PATH, filename)
        self.sheetname = sheetname

    def read_data(self):
        """
        读数据
        :return:
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        testcases_list = []
        headers_list = []  # 存放表头信息
        for row in range(1, ws.max_row + 1):
            # 存放每一行的用例数据
            # one_row_dict = {}
            one_testcase = Testcase()   # 创建用例对象
            for column in range(1, ws.max_column + 1):
                one_cell_value = ws.cell(row, column).value
                if row == 1:
                    # headers_list.append(one_cell_value)
                    # 将获取的表头，转化为字符串并添加至headers_list中
                    headers_list.append(str(one_cell_value))
                else:
                    # 获取表头字符串数据
                    key = headers_list[column - 1]
                    # one_row_dict[key] = one_cell_value
                    if key == "actual":
                        # 设置存放实际响应报文所在列的列号actual_column属性
                        setattr(one_testcase, "actual_column", column)
                    elif key == "result":
                        # 设置存放用例执行结果所在列的列号result_column属性
                        setattr(one_testcase, "result_column", column)

                    setattr(one_testcase, key, one_cell_value)

            if row != 1:
                # testcases_list.append(one_row_dict)
                # 设置当前用例所在的行号row属性
                setattr(one_testcase, "row", row)
                testcases_list.append(one_testcase)

        return testcases_list

    '''
    def write_data(self, row, column, data):
        """
        写操作
        :param row: 指定在某一行写
        :param column: 指定在某一列写
        :param data: 待写入的数据
        :return:
        """
        # 将数据写入到excel中，不能与读取操作公用一个Workbook对象
        # 如果使用同一个Workbook对象，只能将最后一次写入成功，会出现意想不到的结果
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        # 第一种写入方式：
        # one_cell = ws.cell(row, column)
        # one_cell.value = data

        # 第二种写入方式：
        ws.cell(row, column, value=data)

        wb.save(self.filename)
    '''

    def write_data(self, one_testcase, actual_value, result_value):
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        # 第二种写入方式：
        ws.cell(one_testcase.row, one_testcase.actual_column, value=actual_value)
        ws.cell(one_testcase.row, one_testcase.result_column, value=result_value)
        wb.save(self.filename)


if __name__ == '__main__':
    excel_filename = "testcase.xlsx"
    sheet_name = "login"
    do_excel = HandleExcel(excel_filename, sheet_name)
    do_excel.read_data()
    pass

